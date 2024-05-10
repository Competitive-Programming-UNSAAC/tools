import sys
import os
import json
import requests
import openpyxl
import csv
import configparser
from tabulate import tabulate

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import LinearSegmentedColormap
from plottable import ColumnDefinition, Table
from plottable.formatters import decimal_to_percent
from plottable.plots import bar

class User:
    # Basic information
    id = ""
    name = ""
    gender = ""
    codeforcesHandle = ""
    credits = 0
    semester = 0
    contestRegistered = ""
    contestPosition = 0

    # Contest
    isRegisteredOnContest = False

    # Category
    category = ""

    # Codeforces information
    codeforcesRating = 0
    totalRatingSolvedProblems = 0

    # Compute scores
    contestPositionScore = 0.0
    codeforcesRatingScore = 0.0
    ratingSolvedProblemScore = 0.0
    totalScore = 0.0

    def __init__(self, id, name, gender, codeforcesHandle, credits, semester, contestRegistered , contestPosition, codeforcesRating, totalRatingSolvedProblems, useCodeforces):
        self.id = id
        self.name = name
        self.gender = gender
        self.codeforcesHandle = codeforcesHandle
        self.credits = int(credits)
        self.semester = int(semester)
        self.contestRegistered = contestRegistered
        self.contestPosition = int(contestPosition)

        print('Loading Codeforces info for \"{0}\" with handle \"{1}\"'.format(self.name, self.codeforcesHandle))
        if not useCodeforces:
            self.codeforcesRating = int(codeforcesRating)
            self.totalRatingSolvedProblems = int(totalRatingSolvedProblems)
        else:
            self.getCodeforcesRating()
            self.getTotalRatingSolvedProblems()

        self.isRegisteredOnContest()
        print("Loading basic information for \"{0}\"".format(self.name))

    def isRegisteredOnContest(self):
        if self.contestRegistered == "Yes":
            self.isRegisteredOnContest = True

    def getCodeforcesRating(self):
        link = f"https://codeforces.com/api/user.info?handles={self.codeforcesHandle}"
        url = requests.get(link)
        data = json.loads(url.text)

        if 'rating' in data["result"][0].keys():
            self.codeforcesRating = data["result"][0]["rating"]

        # Minimum rating is 800
        self.codeforcesRating = max(800, self.codeforcesRating)


    def getTotalRatingSolvedProblems(self):
        link = f"https://codeforces.com/api/user.status?handle={self.codeforcesHandle}"
        url = requests.get(link)
        data = json.loads(url.text)

        problemSolved = {}
        for submissions in data["result"]:
            # Prepare id for the problem
            key = ""
            if 'problemsetName' in submissions["problem"].keys():
                key += submissions["problem"]["problemsetName"]

            if 'contestId' in submissions["problem"].keys():
                key += str(submissions["problem"]["contestId"])

            if 'index' in submissions["problem"].keys():
                key += submissions["problem"]["index"]    

            # Accumulate problem rating avoid duplicates
            if  key not in problemSolved and submissions['verdict'] == "OK" and "rating" in submissions["problem"]:      
                self.totalRatingSolvedProblems  += submissions["problem"]["rating"]
                problemSolved[key] = True

            # Training contest problems have 1000 of raiting
            if  key not in problemSolved and "rating" not in submissions["problem"]:
                self.totalRatingSolvedProblems += 1000
                problemSolved[key] = True


class Ranking:
    config = {}
    users = []

    totalUsers = 0
    totalRegisteredTeams = 0
    totalContestTeams = 0

    creditsThreshold = 0

    codeforcesRatingAverage = 0.0
    totalRatingSolvedProblemsAverage = 0.0
    maximumCodeforcesRating = 0
    maximumTotalRatingSolvedProblem = 0

    contestPositionWeight = 0
    codeforcesRatingWeight = 0
    totalRatingSolvedProblemsWeight = 0

    def __init__(self, config, users):
        self.config = config
        self.users = users
        self.totalUsers = len(users)
        
        print("Computing ranking...")
        self.computeRanking()
        print("Completed computing the ranking!")

    def getUserCategory(self, user):
        if user.gender == "Female":
            return "W"
        if user.credits == 0:
            if user.semester > 5:
                return "A"
            else:
                return "B"
        if user.credits > self.creditsThreshold:
            return "A"
        else:
            return "B"

    def getContestPositionScore(self, user):
        return (self.totalContestTeams - user.contestPosition + 1) * self.contestPositionWeight / self.totalContestTeams

    def getCodeforcesRatingScore(self, user):
        return (user.codeforcesRating / self.maximumCodeforcesRating) * self.codeforcesRatingWeight

    def getRatingSolvedProblemScore(self, user):
        return (user.totalRatingSolvedProblems / self.maximumTotalRatingSolvedProblem) * self.totalRatingSolvedProblemsWeight

    def computeRanking(self):
        self.totalRegisteredTeams = int(self.config["Contest"]["TotalRegisteredTeams"])
        self.contestPositionWeight = int(self.config["Weight"]["ContestPosition"])
        self.codeforcesRatingWeight = int(self.config["Weight"]["CodeforcesRating"])
        self.totalRatingSolvedProblemsWeight = int(self.config["Weight"]["TotalRatingSolvedProblems"])
        self.creditsThreshold = int(self.config["Credits"]["Threshold"])

        totalCodeforcesRatingAccumulate = 0
        totalRatingSolvedProblemAccumulate = 0
        for user in self.users:
            totalCodeforcesRatingAccumulate += user.codeforcesRating
            totalRatingSolvedProblemAccumulate += user.totalRatingSolvedProblems
            self.maximumCodeforcesRating = max(self.maximumCodeforcesRating, user.codeforcesRating)
            self.maximumTotalRatingSolvedProblem = max(self.maximumTotalRatingSolvedProblem, user.totalRatingSolvedProblems)

            # Count those registered in the contest
            if user.isRegisteredOnContest == True:
                self.totalContestTeams += 1

        self.codeforcesRatingAverage = totalCodeforcesRatingAccumulate / self.totalUsers
        self.totalRatingSolvedProblemsAverage = totalRatingSolvedProblemAccumulate / self.totalUsers

        for user in self.users:
            user.category = self.getUserCategory(user)

            user.contestPositionScore = self.getContestPositionScore(user)
            user.codeforcesRatingScore = self.getCodeforcesRatingScore(user)
            user.ratingSolvedProblemScore = self.getRatingSolvedProblemScore(user)

            user.totalScore = user.contestPositionScore + user.codeforcesRatingScore + user.ratingSolvedProblemScore

        # Sort users by total score
        self.users.sort(key = lambda user : user.totalScore, reverse = True)

    def plotTable(self):
        headers = ["#", "Id", "Name", "Gender", "Handle", "Category", "Contest", "Problems", "Rating", "Contest Score", "Problems Score", "Rating Score", "Total Score"]
        
        table = []
        index = 1
        for user in self.users:
            row = [index, user.id, user.name, user.gender, user.codeforcesHandle, user.category, user.contestPosition , user.totalRatingSolvedProblems, user.codeforcesRating, user.contestPositionScore, user.ratingSolvedProblemScore, user.codeforcesRatingScore, user.totalScore]
            table.append(row)
            index += 1

        rankingTable = tabulate(table, headers=headers, tablefmt='orgtbl', floatfmt=".2f")

        print("Ranking completed!")
        print("")
        print("Contest Position Weight: {0}".format(self.contestPositionWeight))
        print("Codeforces Total Rating of Solved Problems Weight: {0}".format(self.totalRatingSolvedProblemsWeight))
        print("Codeforces Rating Weight: {0}".format(self.codeforcesRatingWeight))
        print("Total participants in the selection: {0}".format(self.totalUsers))
        print("Total registered teams: {0}".format(self.totalRegisteredTeams))
        print("Total contest teams: {0}".format(self.totalContestTeams))
        print("Maximum Codeforces rating of the participants in the selection: {0}".format(self.maximumCodeforcesRating))
        print("Maximum accumulated problem rating of the participants in the selection: {0}".format(self.maximumTotalRatingSolvedProblem))
        print("Average rating of the participants in the selection: {0}".format(self.codeforcesRatingAverage))
        print("Average of the accumulated problem rating of the participants in the selection: {0}".format(self.totalRatingSolvedProblemsAverage))
        print("")
        print(rankingTable)

    def writeFile(self, filepath):
        headers = ["#", "Id", "Name", "Gender", "Handle", "Contest", "Problems", "Rating", "Contest Score", "Problems Score", "Rating Score", "Total Score", "Category", "Attend"]

        data = []
        index = 1
        for user in self.users:
            row = [index, user.id, user.name, user.gender, user.codeforcesHandle, user.contestPosition , user.totalRatingSolvedProblems, user.codeforcesRating, "{0:.2f}".format(user.contestPositionScore), "{0:.2f}".format(user.ratingSolvedProblemScore), "{0:.2f}".format(user.codeforcesRatingScore), "{0:.2f}".format(user.totalScore), user.category, "No"]
            data.append(row)
            index += 1

        filepath = os.path.join(filepath, "ranking.csv")
        print("Writing ranking into \"{0}\" file".format(filepath))
        with open(filepath, 'w', encoding='UTF8', newline='') as file:
            writer = csv.writer(file, delimiter = ",")
            writer.writerow(headers)
            writer.writerows(data)
        print("Ranking file completed!")

    def plotGraphic(self, filepath):
        headers = ["Id", "Name", "Gender", "Handle", "Contest", "Problems", "Rating", "Contest Score %", "Problems Score %", "Rating Score %", "Total Score", "Category"]
        
        data = []
        for user in self.users:
            shortName = user.name
            splitName = user.name.split()
            if len(splitName) > 4:
                shortName = " ".join(splitName[0:3])

            contestPositionScorePercent = user.contestPositionScore / self.contestPositionWeight
            codeforcesRatingScorePercent = user.codeforcesRatingScore / self.codeforcesRatingWeight
            ratingSolvedProblemPercent = user.ratingSolvedProblemScore / self.totalRatingSolvedProblemsWeight

            row = [user.id, shortName, user.gender, user.codeforcesHandle, user.contestPosition , user.totalRatingSolvedProblems, user.codeforcesRating, contestPositionScorePercent, ratingSolvedProblemPercent, codeforcesRatingScorePercent, user.totalScore, user.category]
            data.append(row)
        
        cmap = LinearSegmentedColormap.from_list(
            name="bugw", colors=["#ffffff", "#f2fbd2", "#c9ecb4", "#93d3ab", "#35b0ab"], N=256
        )

        plt.rcParams["font.family"] = ["DejaVu Sans"]
        plt.rcParams["savefig.bbox"] = "tight"

        fig, ax = plt.subplots(figsize=(30, 22))

        dataframe = pd.DataFrame(data, columns = headers)

        table = Table(
            dataframe,
            cell_kw={
                "linewidth": 0,
                "edgecolor": "k",
            },
            textprops={"fontsize": 14, "ha": "center"},
            column_definitions=[
                ColumnDefinition(
                    "Id",
                    width = 0.5,
                ),
                ColumnDefinition(
                    "Name",
                    width = 3.0,
                    textprops = {"fontweight": "bold", "ha": "left"},
                ),
                ColumnDefinition(
                    "Gender",
                    textprops = {"ha": "left"}),
                ColumnDefinition(
                    "Handle",
                    width = 1.2,
                    textprops = {"ha": "left"},
                ),
                ColumnDefinition(
                    "Contest",
                    width = 0.5,
                    textprops = {"fontweight": "bold"},
                ),
                ColumnDefinition(
                    "Problems",
                    textprops = {"fontweight": "bold"},
                ),
                ColumnDefinition(
                    "Rating",
                    textprops = {"fontweight": "bold"},
                ),
                ColumnDefinition(
                    "Contest Score %",
                    width = 1.25,
                    plot_fn = bar,
                    plot_kw = {
                        "cmap": cmap,
                        "plot_bg_bar": True,
                        "annotate": True,
                        "height": 0.5,
                        "lw": 0.5,
                        "formatter": decimal_to_percent,
                    },
                ),
                ColumnDefinition(
                    "Problems Score %",
                    width = 1.25,
                    plot_fn = bar,
                    plot_kw = {
                        "cmap": cmap,
                        "plot_bg_bar": True,
                        "annotate": True,
                        "height": 0.5,
                        "lw": 0.5,
                        "formatter": decimal_to_percent,
                    },
                ),
                ColumnDefinition(
                    "Rating Score %",
                    width = 1.25,
                    plot_fn = bar,
                    plot_kw = {
                        "cmap": cmap,
                        "plot_bg_bar": True,
                        "annotate": True,
                        "height": 0.5,
                        "lw": 0.5,
                        "formatter": decimal_to_percent,
                    },
                ),
                ColumnDefinition(
                    "Total Score",
                    formatter = "{:.2f}",
                    textprops = {"fontweight": "bold"},
                ),
                ColumnDefinition(
                    "Category",
                    width = 0.5,
                ),
            ],
        )

        rowColorsCategory = {
            "A": "#fff9f3",
            "B": "#fffff3",
            "W": "#f3fff8",
        }
        
        selectByCategoryA = int(self.config["Selection"]["CategoryA"])
        selectByCategoryB = int(self.config["Selection"]["CategoryB"])
        selectByCategoryC = int(self.config["Selection"]["CategoryW"])
        numberSelectByCategory = {"A": selectByCategoryA, "B": selectByCategoryB, "W": selectByCategoryC}

        index = 0
        for user in self.users:
            if numberSelectByCategory[user.category] > 0:
                table.rows[index].set_facecolor(rowColorsCategory[user.category])
                numberSelectByCategory[user.category] -= 1
            index += 1

        filepath = os.path.join(filepath, "ranking.png")
        fig.savefig(filepath, facecolor=ax.get_facecolor(), dpi=200)

def readConfig(filepath):
    print("Reading config from \"{0}\" file...".format(filepath))
    config = configparser.ConfigParser()
    config.read(filepath)
    print("Config reading is completed!")
    return config

def readData(filepath):
    print("Reading data from \"{0}\" file...".format(filepath))
    dataframe = openpyxl.Workbook()
    data = dataframe.active

    with open(filepath) as file:
        reader = csv.reader(file, delimiter=',')
        for row in reader:
             data.append(row)

    print("File reading is completed!")
    return data

def getUsers(config, data):
    useCodeforces = int(config["Flag"]["UseCodeforces"])

    idCol = int(config["Column"]["Id"])
    nameCol = int(config["Column"]["Name"])
    genderCol = int(config["Column"]["Gender"])
    codeforcesHandleCol = int(config["Column"]["CodeforcesHandle"])
    creditsCol = int(config["Column"]["Credits"])
    semesterCol = int(config["Column"]["Semester"])
    contestPositionCol = int(config["Column"]["ContestPosition"])
    contestRegisteredCol = int(config["Column"]["ContestRegistered"])

    if not useCodeforces:
        codeforcesRatingCol = int(config["Column"]["CodeforcesRating"])
        totalRatingSolvedProblemsCol = int(config["Column"]["TotalRatingSolvedProblems"])

    print("Loading users information...")
    users = []
    for row in range(2, data.max_row + 1):
        id = data.cell(row, idCol).value
        name = data.cell(row, nameCol).value
        gender = data.cell(row, genderCol).value
        codeforcesHandle = data.cell(row, codeforcesHandleCol).value
        credits = data.cell(row, creditsCol).value
        semester = data.cell(row, semesterCol).value
        contestPosition = data.cell(row, contestPositionCol).value
        contestRegistered = data.cell(row, contestRegisteredCol).value

        codeforcesRating = 0
        totalRatingSolvedProblems = 0
        if not useCodeforces:
            codeforcesRating = data.cell(row, codeforcesRatingCol).value
            totalRatingSolvedProblems = data.cell(row, totalRatingSolvedProblemsCol).value

        if id == None or name == None or gender == None or codeforcesHandle == None or credits == None or semester == None or contestPosition == None or contestRegistered == None:
            continue

        user = User(id, name, gender, codeforcesHandle, credits, semester, contestRegistered, contestPosition, codeforcesRating, totalRatingSolvedProblems, useCodeforces)
        users.append(user)

    print("Loading users information is completed!")
    return users

def main():
    dataFilepath = sys.argv[1]
    outputFilepath = sys.argv[2]
    configFilepath = "Config"

    config = readConfig(configFilepath)
    data = readData(dataFilepath)

    users = getUsers(config, data)
    ranking = Ranking(config, users)

    ranking.plotTable()
    ranking.writeFile(outputFilepath)
    ranking.plotGraphic(outputFilepath)

main()
