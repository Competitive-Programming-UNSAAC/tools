import sys
import os
import openpyxl
import csv
import configparser
import operator

from pathlib import Path

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import LinearSegmentedColormap

from plottable import ColumnDefinition, Table
from plottable.cmap import normed_cmap
from plottable.formatters import decimal_to_percent
from plottable.plots import circled_image # image

class User:
    # Basic information
    id = ""
    name = ""
    codeforcesHandle = ""
    penalty = 0
    problems = []

    # Compute
    totalSolvedProblems = 0

    def __init__(self, id, name, codeforcesHandle, penalty , problems):
        self.id = id
        self.name = name
        self.codeforcesHandle = codeforcesHandle
        self.penalty = int(penalty)
        self.problems = problems
        # Compute total solved problems
        self.getTotalSolvedproblems()

    def getTotalSolvedproblems(self):
        for problem in self.problems:
            if problem == 1:
                self.totalSolvedProblems += 1

class Scoreboard:
    config = {}
    users = []
    totalUsers = 0
    totalProblems = 0

    def __init__(self, config, users):
        self.config = config
        self.users = users
        self.totalUsers = len(users)
        self.totalProblems = int(config["Problem"]["Number"])
        # Sort ranking by total solved problems and penalty
        self.users.sort(key=lambda user:(user.totalSolvedProblems, -user.penalty), reverse = True)

    def plotGraphic(self, filepath):
        data = []
        index = 1
        for user in self.users:
            shortName = user.name
            splitName = user.name.split()
            if len(splitName) > 4:
                shortName = " ".join(splitName[0:3])

            row = [index, user.id, shortName, user.codeforcesHandle, user.totalSolvedProblems, user.penalty]

            for problem in user.problems:
                if problem == 1:
                    row.append(1.0)
                if problem == 0:
                    row.append(0.5)
                if problem == -1:
                    row.append(0.0)

            data.append(row)
            index += 1

        userInfoCol = ["Rank", "Id", "Name", "Handle"]
        contestInfoCol = ["Score", "Penalty"]

        problemsInfoCol = []
        for problemInfo in range(0, self.totalProblems):
            problemsInfoCol.append(chr(ord('A') + problemInfo))

        headers = userInfoCol + contestInfoCol + problemsInfoCol

        df = pd.DataFrame(data, columns=headers)
        df = df.set_index("Rank")
        df.head()

        cmapCell = LinearSegmentedColormap.from_list(
            name="bugw", colors=["#f4958f", "#ffffff", "#35b0ab"], N=256
        )

        cmapText = LinearSegmentedColormap.from_list(
            name="bugw", colors=["#000000", "#ffffff", "#000000"], N=256
        )

        colDefinitions = (
            [
                ColumnDefinition(
                    name="Rank",
                    textprops={"ha": "center", "weight": "bold"},
                    width=1.5,
                ),
                ColumnDefinition(
                    name="Id",
                    textprops={"ha": "center", "weight": "bold"},
                    width=1.5,
                ),
                ColumnDefinition(
                    name="Name",
                    textprops={"ha": "left", "weight": "bold"},
                    width=4.0,
                ),
                ColumnDefinition(
                    name="Handle",
                    textprops={"ha": "left"},
                    width=2.0,
                ),
                ColumnDefinition(
                    name="Score",
                    group="Contest Information",
                    textprops={"ha": "center", "weight": "bold"},
                    width=0.75,
                    cmap=normed_cmap(df["Score"], cmap=matplotlib.cm.Greens),
                ),
                ColumnDefinition(
                    name="Penalty",
                    group="Contest Information",
                    textprops={"ha": "center", "weight": "bold"},
                    width=0.75,
                    cmap=normed_cmap(df["Penalty"], cmap=matplotlib.cm.Reds),
                ),
            ]
            + [
                ColumnDefinition(
                    name=problemsInfoCol[0],
                    title=problemsInfoCol[0].replace(" ", "\n", 1),
                    formatter=decimal_to_percent,
                    cmap=cmapCell,
                    text_cmap = cmapText,
                    group="Problems",
                    border="left",
                    textprops={"ha": "center", "weight": "bold", "bbox": {"boxstyle": "circle", "pad": 0.35}},
                )
            ]
            + [
                ColumnDefinition(
                    name=col,
                    title=col.replace(" ", "\n", 1),
                    formatter=decimal_to_percent,
                    cmap=cmapCell,
                    text_cmap = cmapText,
                    group="Problems",
                    textprops={"ha": "center", "weight": "bold", "bbox": {"boxstyle": "circle", "pad": 0.35}},
                )
                for col in problemsInfoCol[1:]
            ]
        )

        plt.rcParams["font.family"] = ["DejaVu Sans"]
        plt.rcParams["savefig.bbox"] = "tight"

        fig, ax = plt.subplots(figsize=(30, 22))

        table = Table(
            df,
            column_definitions = colDefinitions,
            row_dividers = True,
            footer_divider = True,
            ax = ax,
            textprops = {"fontsize": 14},
            row_divider_kw = {"linewidth": 1, "linestyle": (0, (1, 5))},
            col_label_divider_kw = {"linewidth": 1, "linestyle": "-"},
            column_border_kw = {"linewidth": 1, "linestyle": "-"},
        ).autoset_fontcolors(colnames=["Score", "Penalty"])

        rowColors = {
            "default": "#ececec",
        }

        filepath = os.path.join(filepath, "scoreboard.png")
        fig.savefig(filepath, facecolor=ax.get_facecolor(), dpi=200)

def readConfig(filepath):
    print("Reading config from \"{0}\" file...".format(filepath))
    config = configparser.ConfigParser()
    config.read(filepath)
    print("Config reading is completed!")
    return config

def readData(filepath):
    print("Reading data from \"{0}\" file...".format(filepath))
    dataframe = openpyxl.Workbook()
    data = dataframe.active

    with open(filepath) as file:
        reader = csv.reader(file, delimiter=',')
        for row in reader:
             data.append(row)

    print("File reading is completed!")
    return data

def getUsers(config, data):
    idCol = int(config["Column"]["Id"])
    nameCol = int(config["Column"]["Name"])
    codeforcesHandleCol = int(config["Column"]["CodeforcesHandle"])
    penaltyCol = int(config["Column"]["Penalty"])

    startColumnProblemCol = int(config["Problem"]["StartColumn"])
    numberProblemCol = int(config["Problem"]["Number"])

    print("Loading users information...")
    users = []
    for row in range(2, data.max_row + 1):
        id = data.cell(row, idCol).value
        name = data.cell(row, nameCol).value
        codeforcesHandle = data.cell(row, codeforcesHandleCol).value
        penalty = data.cell(row, penaltyCol).value

        if id == None or name == None or codeforcesHandle == None or penalty == None:
            continue

        problems = []
        for problemCol in range(startColumnProblemCol, startColumnProblemCol + numberProblemCol):
            problemResult = data.cell(row, problemCol).value
            problems.append(int(problemResult))

        user = User(id, name, codeforcesHandle, penalty, problems)
        users.append(user)

    print("Loading users information is completed!")
    return users

def main():
    dataFilepath = sys.argv[1]
    outputFilepath = sys.argv[2]
    configFilepath = "Config"

    config = readConfig(configFilepath)
    data = readData(dataFilepath)

    users = getUsers(config, data)
    scoreboard = Scoreboard(config, users)

    scoreboard.plotGraphic(outputFilepath)

main()
